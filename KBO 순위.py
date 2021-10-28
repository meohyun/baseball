from selenium import webdriver
import datetime
import time
from openpyxl import Workbook   

wb = Workbook()
ws = wb.active
browser = webdriver.Chrome("./chromedriver.exe")

browser.get("https://search.naver.com/search.naver?where=nexearch&sm=top_sug.pre&fbm=1&acr=1&acq=kbo+%EC%88%9C&qdt=0&ie=utf8&query=2021+kbo+%EC%88%9C%EC%9C%84")

kbo_ranks = []
kbo_grades = []
kbo_nums = []
kbo_wins = []
kbo_defs = []
kbo_win_rates = []
kbo_rec_10s = []


for index in range(1,11):
    kbo_rank = browser.find_element_by_xpath(f'//*[@id="teamRankTabPanel_0"]/table/tbody/tr[{index}]/th').text
    kbo_grade = browser.find_element_by_xpath(f'//*[@id="teamRankTabPanel_0"]/table/tbody/tr[{index}]/td[1]/p/span/a').text
    kbo_num = browser.find_element_by_xpath(f'//*[@id="teamRankTabPanel_0"]/table/tbody/tr[{index}]/td[2]').text
    kbo_win = browser.find_element_by_xpath(f'//*[@id="teamRankTabPanel_0"]/table/tbody/tr[{index}]/td[3]').text
    kbo_def = browser.find_element_by_xpath(f'//*[@id="teamRankTabPanel_0"]/table/tbody/tr[{index}]/td[5]').text
    kbo_win_rate =browser.find_element_by_xpath(f'//*[@id="teamRankTabPanel_0"]/table/tbody/tr[{index}]/td[6]').text
    kbo_rec_10 = browser.find_element_by_xpath(f'//*[@id="teamRankTabPanel_0"]/table/tbody/tr[{index}]/td[9]').text
    kbo_ranks.append(kbo_rank)
    kbo_grades.append(kbo_grade)
    kbo_nums.append(kbo_num)
    kbo_wins.append(kbo_win)
    kbo_defs.append(kbo_def)
    kbo_win_rates.append(kbo_win_rate)
    kbo_rec_10s.append(kbo_rec_10) 

dt_now = datetime.datetime.now()
days = ['월','화','수','목','금','토','일']
a = dt_now.weekday()

# 날짜, 시간 정보
local_time = time.localtime()
time_kbo=time.strftime("%Y%m%d",local_time)
start_times = []

print()
print("현재시간: "+ str(dt_now)) 
print('-'*45)
print("순위", "팀명", "경기" , "승", "패" , "승률", "최근10경기")
print()
for i in range(0,10):
    print(kbo_ranks[i] +"위: " + kbo_grades[i] + "/" + kbo_nums[i] + "/" + kbo_wins[i] + "/" + kbo_defs[i] +  "/" +kbo_win_rates[i] + "/" +kbo_rec_10s[i])

# 엑셀파일에 정보 저장
ws.append(("순위", "팀명", "경기" , "승", "패" , "승률", "최근10경기"))
for i in range(0,10):
    ws.append((kbo_ranks[i],kbo_grades[i],kbo_nums[i],kbo_wins[i],kbo_defs[i],kbo_win_rates[i],kbo_rec_10s[i]))


ws["K5"].value ="현재시간"
ws["L5"].value = dt_now
ws.column_dimensions["L"].width = 21
print()
print("일정 생성중...")
print("-"*60)

# KBO 일정
kbo_schedules = []
stadiums = []
browser.find_element_by_xpath('//*[@id="scheduleTab"]').click()
today = browser.find_element_by_xpath(f'//*[@id="cssScheduleSubTab_today_{time_kbo}"]').text
time.sleep(3)


print()
print("날짜: " +today)
print()
print("일정")


# 화 = 0 수 = 1  목 = 2 ,금 =3 토=4 일 =5 화 = 6
for index in range(1,index+1):
    try:
        start_time = browser.find_element_by_xpath(f'//*[@id="myschedule_4"]/table/tbody/tr[{index}]/td[1]').text
        start_times.append(start_time)
        schedule = browser.find_element_by_xpath(f'//*[@id="myschedule_4"]/table/tbody/tr[{index}]/td[2]').text.split('\n')
        kbo_schedules.append(schedule)
        stadium = browser.find_element_by_xpath(f'//*[@id="myschedule_4"]/table/tbody/tr[{index}]/td[3]').text
        stadiums.append(stadium)

    except: 
        continue



for i in range(i):
    try:
        print()
        print(start_times[i],end=" ")
        try:
            for j in range(6):
                print(kbo_schedules[i][j],end=" ") # kbo_schedules 에 indexerror 가 날 경우 
        except IndexError:
            print("경기 취소 되었습니다.", end =" ")  # 경기 취소라고 출력한다.
        print(" 경기장: "+stadiums[i])
    except IndexError:
        break


wb.save("KBO 순위.xlsx")
browser.quit()